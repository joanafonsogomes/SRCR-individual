import pandas as pd
import sys
import openpyxl

def getData(path):
    workbook = openpyxl.load_workbook(path)
    f = open("output.txt","w")
    sheet = workbook.active

    cell = sheet.cell(row=1, column=1)

    for row in range(1,420):
        if not sheet.cell(row=row,column=1).value:
            break

    paragem = []

    for column in range(1,11):
        paragem.append(sheet.cell(row=row,column=column))

getData("dataset-original.xlsx")